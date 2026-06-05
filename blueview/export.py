"""Write markup rows to JSON, CSV, or XLSX."""

from __future__ import annotations

import csv
import json
import sys
from typing import Any


def to_json(rows: list[dict[str, Any]], output_path: str | None = None) -> None:
    text = json.dumps(rows, indent=2, default=str)
    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(text)
    else:
        print(text)


def to_csv(rows: list[dict[str, Any]], output_path: str | None = None) -> None:
    if not rows:
        if output_path:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                pass
        return

    fieldnames = _unified_fieldnames(rows)
    if output_path:
        stream = open(output_path, "w", newline="", encoding="utf-8")
        close_after = True
    else:
        stream = sys.stdout
        close_after = False

    try:
        writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    finally:
        if close_after:
            stream.close()


def to_xlsx(rows: list[dict[str, Any]], output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Markups"

    if not rows:
        wb.save(output_path)
        return

    fieldnames = _unified_fieldnames(rows)

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for col_index, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_index, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_index, row in enumerate(rows, start=2):
        for col_index, name in enumerate(fieldnames, start=1):
            ws.cell(row=row_index, column=col_index, value=row.get(name, ""))

    # Auto-fit column widths (capped at 60)
    for col_index, name in enumerate(fieldnames, start=1):
        max_len = len(name)
        for row in rows:
            val = str(row.get(name, ""))
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_index)].width = min(max_len + 2, 60)

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(output_path)


def to_table(rows: list[dict[str, Any]]) -> None:
    """Print a human-readable fixed-width table to stdout."""
    if not rows:
        print("(no markups)")
        return

    fieldnames = _unified_fieldnames(rows)
    col_widths = {f: len(f) for f in fieldnames}
    for row in rows:
        for f in fieldnames:
            col_widths[f] = max(col_widths[f], len(str(row.get(f, ""))))
    # Cap width so lines don't wrap too badly
    col_widths = {f: min(w, 40) for f, w in col_widths.items()}

    sep = "  "
    header = sep.join(f.ljust(col_widths[f]) for f in fieldnames)
    rule = sep.join("-" * col_widths[f] for f in fieldnames)
    print(header)
    print(rule)
    for row in rows:
        line = sep.join(str(row.get(f, ""))[:col_widths[f]].ljust(col_widths[f]) for f in fieldnames)
        print(line)


def _unified_fieldnames(rows: list[dict[str, Any]]) -> list[str]:
    """Collect all keys across rows, preserving first-seen order."""
    seen: dict[str, None] = {}
    for row in rows:
        for k in row:
            seen[k] = None
    return list(seen.keys())
